"""
fetch_intel_history.py - one-time INTC backtest data runner (like build_universe.py,
run it in a real terminal; the pulls are slow and polite).

Answers: would Altman Z / Piotroski F / Beneish M have flagged Intel's distress
BEFORE the market repriced it in December 2025 / January 2026?

What it does:
  1. Pulls INTC annual AND quarterly statements with their FILING dates from the SEC
     EDGAR companyfacts API (free, no key, proper User-Agent per SEC policy). Falls
     back to yfinance only for anything EDGAR lacks, recording the source per number.
  2. Pulls INTC daily closes 2022-01-01 to 2026-02-28 from yfinance.
  3. Builds point-in-time snapshots with backtest.build_snapshots (which reuses
     data.run_models - no model math is reimplemented anywhere).
  4. Writes:
       data/backtest_intel_raw.json   everything fetched, with filing dates + sources
       data/backtest_intel_trail.csv  one scored row per as-of date
       backtest/INTC-backtest.xlsx    Trail / Sources / Read workbook (openpyxl)

Resumable: raw pulls are cached under data/backtest_cache/; delete that folder to
force a refetch. Every network call is retried once, never hammered.

Run:  python3 fetch_intel_history.py
"""
from __future__ import annotations

import csv
import datetime as dt
import json
import os
import time

import requests

from backtest import build_snapshots, classify_trail, EVENT_START, EVENT_END

CIK = "0000050863"                      # Intel Corporation
TICKER = "INTC"
UA = {"User-Agent": "Michael Spring mspring823@gmail.com (financial-health-screener backtest)"}
CACHE_DIR = "data/backtest_cache"
RAW_PATH = "data/backtest_intel_raw.json"
CSV_PATH = "data/backtest_intel_trail.csv"
XLSX_PATH = "backtest/INTC-backtest.xlsx"
PRICE_START, PRICE_END = "2022-01-01", "2026-02-28"

# LINE_ITEMS (data.py) -> candidate us-gaap tags, first hit wins. "flow" items are
# durations (income/cash-flow statement); the rest are instants (balance sheet).
TAGS = {
    "sales": ["RevenueFromContractWithCustomerExcludingAssessedTax", "Revenues"],
    "cogs": ["CostOfGoodsAndServicesSold", "CostOfRevenue", "CostOfGoodsSold"],
    "depreciation": ["DepreciationDepletionAndAmortization",
                     "DepreciationAmortizationAndAccretionNet", "Depreciation"],
    "sga": ["SellingGeneralAndAdministrativeExpense",
            "MarketingGeneralAndAdministrativeExpense",
            "GeneralAndAdministrativeExpense"],
    "net_income": ["NetIncomeLoss"],
    "cfo": ["NetCashProvidedByUsedInOperatingActivities",
            "NetCashProvidedByUsedInOperatingActivitiesContinuingOperations"],
    "ebit": ["OperatingIncomeLoss"],
    "receivables": ["AccountsReceivableNetCurrent", "ReceivablesNetCurrent"],
    "current_assets": ["AssetsCurrent"],
    "current_liabilities": ["LiabilitiesCurrent"],
    "ppe": ["PropertyPlantAndEquipmentNet",
            "PropertyPlantAndEquipmentAndFinanceLeaseRightOfUseAssetAfterAccumulatedDepreciationAndAmortization"],
    "total_assets": ["Assets"],
    "long_term_debt": ["LongTermDebtNoncurrent", "LongTermDebt"],
    "retained_earnings": ["RetainedEarningsAccumulatedDeficit"],
    "total_liabilities": ["Liabilities"],
}
# Some filers (Intel included) never tag "Liabilities"; it is derived per side as
# total assets minus stockholders' equity, the same fallback data.fetch_live uses.
EQUITY_TAGS = ["StockholdersEquity",
               "StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest"]
FLOW_ITEMS = ("sales", "cogs", "depreciation", "sga", "net_income", "cfo", "ebit")
INSTANT_ITEMS = tuple(k for k in TAGS if k not in FLOW_ITEMS)


# ----------------------------------------------------------------------------
# Cached, polite fetching
# ----------------------------------------------------------------------------
def _cached_json(path, fetch):
    if os.path.exists(path):
        with open(path) as fh:
            return json.load(fh)
    data = fetch()
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w") as fh:
        json.dump(data, fh)
    return data


def fetch_companyfacts() -> dict:
    def pull():
        url = f"https://data.sec.gov/api/xbrl/companyfacts/CIK{CIK}.json"
        for attempt in (1, 2):
            try:
                r = requests.get(url, headers=UA, timeout=60)
                r.raise_for_status()
                return r.json()
            except Exception as e:  # noqa: BLE001
                if attempt == 2:
                    raise
                print(f"  EDGAR pull failed ({e}), retrying once in 5s...")
                time.sleep(5)
    return _cached_json(f"{CACHE_DIR}/companyfacts_{TICKER}.json", pull)


def fetch_prices() -> dict:
    def pull():
        import yfinance as yf
        for attempt in (1, 2):
            try:
                hist = yf.Ticker(TICKER).history(start=PRICE_START, end=PRICE_END,
                                                 auto_adjust=False)
                closes = {d.strftime("%Y-%m-%d"): float(v)
                          for d, v in hist["Close"].items() if v == v}
                if not closes:
                    raise RuntimeError("empty price history")
                return closes
            except Exception as e:  # noqa: BLE001
                if attempt == 2:
                    raise
                print(f"  price pull failed ({e}), retrying once in 5s...")
                time.sleep(5)
    return _cached_json(f"{CACHE_DIR}/prices_{TICKER}.json", pull)


# ----------------------------------------------------------------------------
# EDGAR fact extraction
# ----------------------------------------------------------------------------
def _facts_for(facts: dict, item: str, tags=None):
    """
    All USD facts across the item's candidate tags, MERGED (a filer can switch tags
    over time, e.g. Intel's PP&E). The indexers keep the earliest-filed fact per
    period, so overlaps resolve deterministically. Returns (joined tag names, facts).
    """
    gaap = facts.get("facts", {}).get("us-gaap", {})
    used, out = [], []
    for tag in (tags or TAGS[item]):
        node = gaap.get(tag)
        if not node:
            continue
        units = node.get("units", {})
        vals = units.get("USD") or next(iter(units.values()), [])
        if vals:
            used.append(tag)
            out.extend(vals)
    return ("|".join(used) or None), out


def _days(a: str, b: str) -> int:
    return (dt.date.fromisoformat(b) - dt.date.fromisoformat(a)).days


def _index_durations(vals):
    """{(start, end): fact} keeping the earliest-filed fact per duration."""
    out = {}
    for v in vals:
        if "start" not in v or v.get("val") is None:
            continue
        key = (v["start"], v["end"])
        if key not in out or v["filed"] < out[key]["filed"]:
            out[key] = v
    return out


def _index_instants(vals):
    """{end: fact} keeping the earliest-filed fact per instant date."""
    out = {}
    for v in vals:
        if v.get("val") is None:
            continue
        end = v["end"]
        if end not in out or v["filed"] < out[end]["filed"]:
            out[end] = v
    return out


def _near(target: str, candidates, tol=15):
    """The candidate date string closest to target within tol days, else None."""
    t = dt.date.fromisoformat(target)
    best = None
    for c in candidates:
        gap = abs((dt.date.fromisoformat(c) - t).days)
        if gap <= tol and (best is None or gap < best[1]):
            best = (c, gap)
    return best[0] if best else None


class FactBook:
    """Per-item duration/instant indexes over the whole companyfacts payload."""

    def __init__(self, facts: dict):
        self.tag = {}
        self.dur = {}       # item -> {(start,end): fact}
        self.inst = {}      # item -> {end: fact}
        for item in FLOW_ITEMS:
            tag, vals = _facts_for(facts, item)
            self.tag[item] = tag
            self.dur[item] = _index_durations(vals)
        for item in INSTANT_ITEMS:
            tag, vals = _facts_for(facts, item)
            self.tag[item] = tag
            self.inst[item] = _index_instants(vals)
        # equity instants, for deriving total liabilities when "Liabilities" is untagged
        _, eq_vals = _facts_for(facts, "total_liabilities", tags=EQUITY_TAGS)
        self.equity = _index_instants(eq_vals)
        # shares outstanding from the dei cover facts: {filed: val}
        dei = facts.get("facts", {}).get("dei", {}).get(
            "EntityCommonStockSharesOutstanding", {})
        self.shares_by_filed = {}
        for v in next(iter(dei.get("units", {}).values()), []):
            if v.get("val"):
                self.shares_by_filed[v["filed"]] = float(v["val"])

    # ---- annual / fiscal-year structure ----
    def fy_ends(self, item="total_assets"):
        """All annual duration end dates, from net_income (the most reliably tagged)."""
        return sorted(end for (start, end) in self.dur["net_income"]
                      if 340 <= _days(start, end) <= 380)

    def annual(self, item: str, end: str):
        """The ~1-year duration ending at `end`, else None."""
        for (s, e), f in self.dur[item].items():
            if e == end and 340 <= _days(s, e) <= 380:
                return f
        return None

    def ytd(self, item: str, end: str):
        """The longest sub-annual duration ending at `end` (the fiscal YTD figure)."""
        best = None
        for (s, e), f in self.dur[item].items():
            n = _days(s, e)
            if e == end and n < 340 and (best is None or n > best[0]):
                best = (n, f)
        return best[1] if best else None

    def ttm(self, item: str, q_end: str, fy_ends):
        """
        Trailing-twelve-month flow ending at quarter end q_end:
          TTM = FY(prior fiscal year) + YTD(q_end) - YTD(same quarter one year earlier)
        Returns (value, note) or (None, reason).
        """
        prior_fy = max((e for e in fy_ends if e < q_end), default=None)
        if prior_fy is None:
            return None, "no prior fiscal year on file"
        fy = self.annual(item, prior_fy)
        cur = self.ytd(item, q_end)
        prior_q = _near((dt.date.fromisoformat(q_end)
                         - dt.timedelta(days=364)).isoformat(),
                        {e for (_s, e) in self.dur[item]})
        pri = self.ytd(item, prior_q) if prior_q else None
        if fy is None or cur is None or pri is None:
            return None, "missing FY or YTD durations for TTM"
        return fy["val"] + cur["val"] - pri["val"], f"TTM = FY {prior_fy} + YTD - prior YTD"

    def instant(self, item: str, end: str):
        f = self.inst[item].get(end)
        if f is None:
            near = _near(end, self.inst[item].keys(), tol=7)
            f = self.inst[item].get(near) if near else None
        return f

    def equity_at(self, end: str):
        f = self.equity.get(end)
        if f is None:
            near = _near(end, self.equity.keys(), tol=7)
            f = self.equity.get(near) if near else None
        return f

    def shares_asof(self, filed: str):
        """Shares outstanding from the cover of the most recent filing at `filed`."""
        cands = [d for d in self.shares_by_filed if d <= filed]
        return self.shares_by_filed[max(cands)] if cands else None

    def filed_date(self, form: str, period_end: str):
        """Earliest filed date among facts of `form` for this period (the filing date)."""
        dates = []
        for item in INSTANT_ITEMS:
            for v in [f for e, f in self.inst[item].items() if e == period_end]:
                if v.get("form") == form:
                    dates.append(v["filed"])
        for item in FLOW_ITEMS:
            for (s, e), v in self.dur[item].items():
                if e == period_end and v.get("form") == form:
                    dates.append(v["filed"])
        return min(dates) if dates else None


# ----------------------------------------------------------------------------
# Filing assembly (payload-ready dicts for backtest.build_snapshots)
# ----------------------------------------------------------------------------
def _yf_annual_fallback():
    """Annual statement values from yfinance keyed by period year, for EDGAR gaps."""
    try:
        import yfinance as yf
        from data import _row
        t = yf.Ticker(TICKER)
        bs, ic, cf = t.balance_sheet, t.income_stmt, t.cashflow
        out = {}
        frames = {"sales": (ic, ["Total Revenue"]), "cogs": (ic, ["Cost Of Revenue"]),
                  "ebit": (ic, ["EBIT", "Operating Income"]),
                  "sga": (ic, ["Selling General And Administration"]),
                  "net_income": (ic, ["Net Income"]),
                  "cfo": (cf, ["Operating Cash Flow"]),
                  "depreciation": (cf, ["Depreciation And Amortization"])}
        for df in (bs, ic, cf):
            if df is None or getattr(df, "empty", True):
                return {}
        for col_i, col in enumerate(ic.columns):
            year = str(col)[:4]
            vals = {}
            for item, (df, labels) in frames.items():
                pair = _row(df, *labels)
                vals[item] = pair[0] if col_i == 0 else (pair[1] if col_i == 1 else None)
            out[year] = vals
        return out
    except Exception as e:  # noqa: BLE001
        print(f"  yfinance fallback unavailable ({e}); EDGAR-only run.")
        return {}


def build_filings(book: FactBook, sources: list):
    """
    One payload-ready filing per 10-K (annual pair) and per 10-Q (TTM flows + quarter
    balance sheet vs the same view one year earlier). Missing items become None and
    the affected model degrades inside run_models, same as the live app.
    """
    filings = []
    fy_ends = book.fy_ends()
    yf_fb = None                          # lazily loaded only if EDGAR has gaps

    def rec(period, form, side, item, val, tag, filed, note=""):
        sources.append({"period_end": period, "form": form, "side": side, "item": item,
                        "value": val, "tag": tag or "", "filed": filed or "",
                        "source": note or "EDGAR companyfacts"})

    def derive_tl(d, pe, period, form, side):
        """Total liabilities = assets - equity when the filer never tags Liabilities."""
        if d.get("total_liabilities") is None and pe and d.get("total_assets") is not None:
            eq = book.equity_at(pe)
            if eq:
                d["total_liabilities"] = d["total_assets"] - eq["val"]
                rec(period, form, side, "total_liabilities", d["total_liabilities"],
                    "derived: Assets - StockholdersEquity", eq["filed"],
                    "derived from EDGAR equity")

    def fallback(item, period_end):
        nonlocal yf_fb
        if item not in FLOW_ITEMS:
            return None
        if yf_fb is None:
            yf_fb = _yf_annual_fallback()
        v = (yf_fb.get(period_end[:4]) or {}).get(item)
        return v

    # ---- annual filings (10-K): curr = FY, prior = previous FY ----
    for i, end in enumerate(fy_ends):
        filed = book.filed_date("10-K", end)
        if filed is None:
            continue
        prior_end = fy_ends[i - 1] if i > 0 else None
        curr, prior = {}, {}
        for item in FLOW_ITEMS:
            for side, pe in (("curr", end), ("prior", prior_end)):
                f = book.annual(item, pe) if pe else None
                val = f["val"] if f else (fallback(item, pe) if pe else None)
                (curr if side == "curr" else prior)[item] = val
                rec(end, "10-K", side, item, val, book.tag[item],
                    f["filed"] if f else "",
                    "" if f else ("yfinance fallback" if val is not None else "missing"))
        for item in INSTANT_ITEMS:
            for side, pe in (("curr", end), ("prior", prior_end)):
                f = book.instant(item, pe) if pe else None
                val = f["val"] if f else None
                (curr if side == "curr" else prior)[item] = val
                rec(end, "10-K", side, item, val, book.tag[item],
                    f["filed"] if f else "", "" if f else "missing")
        derive_tl(curr, end, end, "10-K", "curr")
        derive_tl(prior, prior_end, end, "10-K", "prior")
        curr["shares"] = prior["shares"] = book.shares_asof(filed)
        rec(end, "10-K", "curr", "shares", curr["shares"],
            "dei:EntityCommonStockSharesOutstanding", filed)
        filings.append({"filing_date": filed, "period_end": end, "form": "10-K",
                        "curr": curr, "prior": prior, "shares": curr["shares"]})

    # ---- quarterly filings (10-Q): TTM flows, quarter-end balance sheet ----
    q_ends = sorted({e for e in book.inst["total_assets"]
                     if book.inst["total_assets"][e].get("form") == "10-Q"})
    for q in q_ends:
        filed = book.filed_date("10-Q", q)
        if filed is None:
            continue
        prior_q = _near((dt.date.fromisoformat(q) - dt.timedelta(days=364)).isoformat(),
                        set(book.inst["total_assets"].keys()))
        curr, prior = {}, {}
        for item in FLOW_ITEMS:
            val, note = book.ttm(item, q, fy_ends)
            curr[item] = val
            rec(q, "10-Q", "curr", item, val, book.tag[item], filed,
                note if val is not None else f"missing ({note})")
            pval, pnote = book.ttm(item, prior_q, fy_ends) if prior_q else (None, "no prior quarter")
            prior[item] = pval
            rec(q, "10-Q", "prior", item, pval, book.tag[item], "",
                pnote if pval is not None else f"missing ({pnote})")
        for item in INSTANT_ITEMS:
            for side, pe in (("curr", q), ("prior", prior_q)):
                f = book.instant(item, pe) if pe else None
                val = f["val"] if f else None
                (curr if side == "curr" else prior)[item] = val
                rec(q, "10-Q", side, item, val, book.tag[item],
                    f["filed"] if f else "", "" if f else "missing")
        derive_tl(curr, q, q, "10-Q", "curr")
        derive_tl(prior, prior_q, q, "10-Q", "prior")
        curr["shares"] = prior["shares"] = book.shares_asof(filed)
        rec(q, "10-Q", "curr", "shares", curr["shares"],
            "dei:EntityCommonStockSharesOutstanding", filed)
        filings.append({"filing_date": filed, "period_end": q, "form": "10-Q",
                        "curr": curr, "prior": prior, "shares": curr["shares"]})

    filings.sort(key=lambda f: (f["filing_date"], f["period_end"]))
    return filings


# ----------------------------------------------------------------------------
# As-of dates: quarterly steps 2022-06-30 .. 2026-01-31, plus the latest price date
# ----------------------------------------------------------------------------
def as_of_dates(prices: dict):
    out, d = [], dt.date(2022, 6, 30)
    while d <= dt.date(2025, 12, 31):
        out.append(d.isoformat())
        # step to the next calendar quarter end
        month = {3: (3, 31), 6: (6, 30), 9: (9, 30), 12: (12, 31)}
        q = min((m for m in (3, 6, 9, 12) if m > d.month), default=None)
        d = dt.date(d.year, q, month[q][1]) if q else dt.date(d.year + 1, 3, 31)
    out.append("2026-01-31")
    latest = max(prices.keys())
    if latest not in out:
        out.append(latest)
    return sorted(out)


# ----------------------------------------------------------------------------
# Excel workbook (Trail / Sources / Read)
# ----------------------------------------------------------------------------
def write_workbook(rows, sources, verdict):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    GREEN = PatternFill("solid", fgColor="C7EFCF")
    AMBER = PatternFill("solid", fgColor="FCE8B2")
    RED = PatternFill("solid", fgColor="F8C8CE")
    BOLD = Font(bold=True)

    wb = Workbook()

    # ---- Sheet 1: Trail ----
    ws = wb.active
    ws.title = "Trail"
    cols = ["as_of_date", "statement_period_used", "filing_date_used", "form_used",
            "z", "zone", "f_score", "m_score", "m_flag", "verdict", "integrity",
            "price_used", "mve_used", "note"]
    ws.append(cols)
    for c in ws[1]:
        c.font = BOLD
    for r in rows:
        ws.append([r.get(k) for k in cols])
        line = ws.max_row
        zone = r.get("zone")
        if zone:
            ws.cell(line, cols.index("zone") + 1).fill = (
                GREEN if zone == "Safe" else AMBER if zone == "Grey" else RED)
        f = r.get("f_score")
        if f is not None:
            ws.cell(line, cols.index("f_score") + 1).fill = (
                GREEN if f >= 7 else AMBER if f >= 4 else RED)
        if r.get("m_flag") is not None:
            ws.cell(line, cols.index("m_flag") + 1).fill = RED if r["m_flag"] else GREEN
        v = r.get("verdict")
        if v in ("Healthy", "Watch", "Distressed"):
            ws.cell(line, cols.index("verdict") + 1).fill = (
                GREEN if v == "Healthy" else AMBER if v == "Watch" else RED)
    for i, w in enumerate([11, 19, 15, 9, 7, 9, 8, 9, 8, 11, 20, 10, 16, 40], 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    # ---- Sheet 2: Sources ----
    ws2 = wb.create_sheet("Sources")
    scols = ["period_end", "form", "side", "item", "value", "tag", "filed", "source"]
    ws2.append(scols)
    for c in ws2[1]:
        c.font = BOLD
    for s in sources:
        ws2.append([s.get(k) for k in scols])
    for i, w in enumerate([12, 7, 7, 18, 16, 44, 12, 34], 1):
        ws2.column_dimensions[ws2.cell(1, i).column_letter].width = w

    # ---- Sheet 3: Read ----
    ws3 = wb.create_sheet("Read")
    ws3["A1"] = "Would the models have flagged Intel before December 2025?"
    ws3["A1"].font = BOLD
    lines = [
        verdict["summary"],
        "",
        "Method: at each as-of date, only filings whose FILING date is on or before that "
        "date are used, and market value of equity is the close on that date times shares "
        "outstanding from the selected filing. No number here uses information from after "
        "its own row's date.",
        "Warning states: Altman Z in the grey or distress zone, Piotroski F of 3 or below, "
        "or a Beneish M manipulation flag.",
        f"Event window: {verdict['event_start']} to {verdict['event_end']}. "
        f"First warning: {verdict['first_warning_date'] or 'never'}"
        + (f", {verdict['lead_days']} days of lead time." if verdict["lead_days"] else "."),
        "Scores are computed from public filings. Educational analysis, not investment advice.",
    ]
    for i, line in enumerate(lines, 3):
        ws3.cell(i, 1, line)
    ws3.column_dimensions["A"].width = 110

    os.makedirs(os.path.dirname(XLSX_PATH), exist_ok=True)
    wb.save(XLSX_PATH)


# ----------------------------------------------------------------------------
def main():
    print("1/5 Fetching EDGAR companyfacts for INTC (cached after first run)...")
    facts = fetch_companyfacts()
    print("2/5 Fetching INTC daily prices (cached after first run)...")
    prices = fetch_prices()

    print("3/5 Assembling point-in-time filings...")
    book = FactBook(facts)
    sources = []
    filings = build_filings(book, sources)
    print(f"    {len(filings)} filings assembled "
          f"({sum(1 for f in filings if f['form'] == '10-K')} 10-K, "
          f"{sum(1 for f in filings if f['form'] == '10-Q')} 10-Q)")

    print("4/5 Scoring the trail...")
    dates = as_of_dates(prices)
    rows = build_snapshots(filings, prices, dates)
    verdict = classify_trail(rows, EVENT_START, EVENT_END)

    print("5/5 Writing outputs...")
    os.makedirs("data", exist_ok=True)
    with open(RAW_PATH, "w") as fh:
        json.dump({"ticker": TICKER, "cik": CIK,
                   "fetched": sorted(os.listdir(CACHE_DIR)),
                   "filings": filings, "sources": sources,
                   "prices": prices, "as_of_dates": dates,
                   "verdict": verdict}, fh, indent=1)
    with open(CSV_PATH, "w", newline="") as fh:
        cols = ["as_of_date", "statement_period_used", "filing_date_used", "form_used",
                "z", "zone", "f_score", "m_score", "m_flag", "verdict", "integrity",
                "price_used", "mve_used", "note"]
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k) for k in cols})
    write_workbook(rows, sources, verdict)

    print(f"\nDone. {RAW_PATH}, {CSV_PATH}, {XLSX_PATH}")
    print("\n" + verdict["summary"])


if __name__ == "__main__":
    main()
